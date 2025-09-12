import openpyxl
from openpyxl.utils import get_column_letter
from django.utils import timezone
from django.http import HttpResponse
from apps.transactions.models import Transaction

def export_as_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions of the Month"

    # Header
    headers = [
        'Date', 'Type', 'Name', 'Price', 'Category', 'User',
    ]
    ws.append(headers)

    # Loop รายงานแต่ละรายการ เพื่อดึง transactions ทั้งหมดในช่วงนั้น
    for report in queryset:
        # ดึง transaction ตามช่วงเวลา + user
        transactions = Transaction.objects.filter(
            user=report.user,
            date__date__gte=report.start_date,
            date__date__lte=report.end_date
        ).select_related('category')

        for tx in transactions:
            tx_date = tx.date.astimezone(timezone.get_current_timezone()).replace(tzinfo=None)


            ws.append([
                tx_date,
                tx.get_type_display(),
                tx.name,
                float(tx.price),
                tx.category.name if tx.category else '—',
                tx.user.username,
            ])

    # Adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[get_column_letter(column)].width = max_length + 2

    # Return response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=monthly_transactions.xlsx'
    wb.save(response)
    return response